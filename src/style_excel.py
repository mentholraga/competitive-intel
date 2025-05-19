# src/style_excel.py

import json
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side

def flatten_dict(d: dict, parent_key: str = "") -> dict:
    """
    Recursively flattens a nested dict by joining keys with ' — '.
    """
    items = {}
    for k, v in d.items():
        new_key = f"{parent_key}{k}" if not parent_key else f"{parent_key} — {k}"
        if isinstance(v, dict):
            items.update(flatten_dict(v, new_key))
        else:
            items[new_key] = v
    return items

def style_excel(company: str):
    # 1) Load your cleaned JSON
    json_path = f"data/output_{company}.json"
    with open(json_path, "r") as f:
        data = json.load(f)

    # 2) Unwrap a single‐root if needed
    if len(data) == 1 and isinstance(next(iter(data.values())), dict):
        data = next(iter(data.values()))

    # 3) Flatten the nested dict
    flat = flatten_dict(data)

    # 4) Create a new workbook and write header + rows
    wb = Workbook()
    ws = wb.active
    ws.title = "Competitive Intel"

    ws.append(["Field", "Value"])
    for field, value in flat.items():
        ws.append([field, value])

    # 5) Style header row
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True)

    # 6) Set reasonable column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 80

    # 7) Draw thin borders around every cell
    thin = Side(border_style='thin')
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # 8) Freeze the header
    ws.freeze_panes = 'A2'

    # 9) Save the styled workbook
    out_path = f"data/output_{company}_styled.xlsx"
    wb.save(out_path)
    print(f"✅ Wrote styled Excel to {out_path}")

if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("Usage: python src/style_excel.py <company_name>")
        sys.exit(1)
    style_excel(sys.argv[1])
